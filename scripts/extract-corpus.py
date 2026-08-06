# Fusionne les fichiers sources de correspondance VIN (xlsx + xls) en un corpus
# unique dédupliqué par VIN → scripts/echantillon.csv (colonnes
# vin,marque,modele,annee,vehicule,type,finition). Régénérer ensuite le seed via
# gen-vin-seed.py. Dépendances : openpyxl (.xlsx) + xlrd (.xls).
# ⚠️ Chemins sources = dossier de données de l'utilisateur (hors dépôt).
import openpyxl, xlrd, csv, sys
sys.stdout.reconfigure(encoding='utf-8')  # console Windows (cp1252) → UTF-8

DOSSIER = r'F:\AI PROJECTS\Table correspondanceVIN'
SOURCES = [
    ('xlsx', DOSSIER + r'\ExportCorespondance VIN_Marque_Modele.xlsx'),
    ('xls',  DOSSIER + r'\Export VIN du 01_01_2010 au 31_12_2020.xls'),
]
OUT = r'F:\AI PROJECTS\STCA-Electron\scripts\echantillon.csv'
# Indices colonnes (identiques dans les deux fichiers) :
C_VIN, C_MARQUE, C_MODELE, C_TYPE, C_FINITION, C_ANNEE, C_VEHIC = 0, 1, 2, 3, 4, 5, 6

def annee(v):
    try: return str(int(float(v)))
    except Exception: return ''

def s(v):
    return '' if v is None else str(v).strip()

seen = set()
rows = []
stats = {}

def ajoute(src, r):
    vin = s(r[C_VIN]).upper().replace(' ', '')
    if not vin or vin in seen:
        stats[src] = stats.get(src, [0, 0]); stats[src][1] += 1
        return
    seen.add(vin)
    rows.append([vin, s(r[C_MARQUE]), s(r[C_MODELE]), annee(r[C_ANNEE]), s(r[C_VEHIC]), s(r[C_TYPE]), s(r[C_FINITION])])
    stats[src] = stats.get(src, [0, 0]); stats[src][0] += 1

for kind, path in SOURCES:
    if kind == 'xlsx':
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]; it = ws.iter_rows(values_only=True); next(it)
        for r in it: ajoute(path.split('\\')[-1], r)
    else:
        xb = xlrd.open_workbook(path); xs = xb.sheet_by_index(0)
        for ri in range(1, xs.nrows):
            ajoute(path.split('\\')[-1], [xs.cell_value(ri, c) for c in range(xs.ncols)])

with open(OUT, 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f); w.writerow(['vin', 'marque', 'modele', 'annee', 'vehicule', 'type', 'finition']); w.writerows(rows)

print('CORPUS COMBINÉ →', OUT)
for src, (gardes, doublons) in stats.items():
    print(f'  {src[:45]:45} : +{gardes} gardés, {doublons} doublons/vides ignorés')
print('  TOTAL lignes uniques :', len(rows))
